import openpyxl, os 
from openpyxl import Workbook
import datetime
import time 
import getpass 

def count_time(start):
    time_sec = start
    time_hour, time_sec =  time_sec // 3600, time_sec % 3600
    time_min, time_sec = time_sec // 60, time_sec % 60
    time_hour = str(int(time_hour))
    time_min = str(int(time_min))
    time_sec = str(round(time_sec, 2))
    time_sec = time_sec.split('.')
    time_mil = time_sec[1]
    time_sec = time_sec[0]
    count_up_str = "{}:{}:{}.{}".format(time_hour.zfill(2),
                                                    time_min.zfill(2),
                                                    time_sec.zfill(2),
                                                    time_mil.zfill(2))
    return count_up_str

def secs2time(s):
    ms = int((s - int(s)) * 1000000)
    s = int(s)
    # Get rid of this line if s will never exceed 86400
    #while s >= 24*60*60:
        #s -= 24*60*60
    h = s / (60*60)
    s -= h*60*60
    m = s / 60
    s -= m*60
    return datetime.time(h, m, s, ms)

# function to get the required time for each distance
def time_calc(dist, tspeed):
    hour_in_seconds = 3600.0
    return round((dist / tspeed) * hour_in_seconds, 1)

def get_time_list(total, speed):
    current_dist = 0.0
    time_list = []

    #while (current_dist < total - .1):
    for i in range(0, total):
        current_dist += .1
        #time_list.append(str(secs2time(time_calc(current_dist, speed))))
        time_list.append(count_time(time_calc(current_dist, speed)))
    
    return time_list

def create_folder(path):
    try:
        os.mkdir(path)
    except OSError:
        print ("Creation of the directory %s failed" % path)
    else:
        print ("Successfully created the directory %s " % path)    

def save_list_in_excel(pulse_list, target_list, mile_list, tire_size, target_miles, target_speed, number_of_legs, end_1st_leg=0):
    # Create the spreadsheet 
    # filename = r"C:\Users\AIOTIK-005\Desktop\Route_created.xlsx"
    book = Workbook()
    ws = book.active

    # number of legs
    ws['D1'] = "Number of Legs"
    ws['D2'] = int(number_of_legs)
    ws['E1'] = "End of 1st Leg"
    ws['E2'] = int(end_1st_leg)

    # Reference
    ws['H1'] = "REF"

    # write the target data here
    ws['I1'] = target_speed
    ws['J1'] = "MPH"

    ws['K1'] = tire_size
    ws['L1'] = "Inch"
    
    ws['M1'] = target_miles
    ws['N1'] = "Miles"

    # setup colum A and B width
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    # set sheet name 
    ws.append(["PULSES", "TIME", "MILES"])

    # append the values to the column 
    # append the mile_list to the spreadsheet 
    r = 3
    for statN in mile_list:
        ws.cell(row = r, column = 3).value = statN
        r += 1
    
    # append the pulse list to the spreadsheet
    r = 3
    for statN in pulse_list:
        # ws.cell(row = r, column = 1).value = '=(63360/K1/2) * {}'.format(str(r - 2))
        ws.cell(row = r, column = 1).value = '{}'.format(statN)
        ws.cell(row = r, column = 1).number_format = '0'
        #ws.cell(row = r, column = 1).value = statN
        r += 1
    
    # append the target list to spreadsheet
    r = 3
    for statN in target_list:
        ws.cell(row = r, column = 2).value = '=((C{} / I1) * 3600.0) / 86400'.format(r)
        ws.cell(row = r, column = 2).number_format = 'hh:mm:ss.00'
        r += 1

    # create the spreadsheet 
    user = os.getcwd().split('\\')
    user = user[0] + '\\' + user[1] + '\\' + user[2]
    filename = str(input("Enter the name for the spreadsheet: "))
    filepath = user + "\Desktop\\" + filename + r'.xlsx'
    # filepath = r"C:\Users\\" + getpass.getuser() +  r"\Desktop\\" + filename + r".xlsx"
    folder_name = 'race_data'
    create_folder(folder_name)
    path = os.getcwd() + '\\' + folder_name + '\\' + filename + r'.xlsx'
    print(f"Your file location is here: {path}")
    book.save(path)

def calculate_pulses(tire_size):
    return (63360 / tire_size) / 2         # rotations per mile = (mile in inches / tire circumference) 
                                           # 6336 = 0.1 miles
                                           # 528 feet = 6336 inches

def get_pulse_by_distance(first_pulse, distance_end_1st_leg):
    distance_list = []
    mile_counter = 0
    while (mile_counter < distance_end_1st_leg):
        mile_counter += .1
        mile_counter = round(mile_counter, 1)
        distance_list.append(mile_counter)

    dist_result = round(first_pulse * len(distance_list), 0)
    return dist_result

def get_pulses_list(first_pulse, total):
    multiplier = 0
    pulse_list = []

    for i in range(0, total):
        multiplier += 1
        result = round(first_pulse * multiplier, 0)
        pulse_list.append(result)
    
    return pulse_list

def get_miles_list(total):
    mile_list = []
    mile_counter = 0

    while (mile_counter < total):
        mile_counter += .1
        mile_counter = round(mile_counter, 1)
        mile_list.append(mile_counter)
    
    return mile_list

def convert_miles_to_feet(miles):
    return (miles - .1) * 5280

def convert_feet_to_inch(feet):
    return feet * 12

def calculate_tire_rotation_from_feet(inches, tire_size):
    return round(inches / tire_size, 2)

def calculate_total_pulses(tire_rotation):
    return tire_rotation * 5

def calculate_minimum_pulse(total_pulses, distance_size):
    return total_pulses / distance_size

def calculate_list_pulses(total_pulses, distance_size, more_pulses=0):
    pulse_list = []
    for index in range(distance_size + 1):
        # if index != 0:
        conversion = round((total_pulses * index) + more_pulses, 2)
        pulse_list.append(conversion)
        # print(f"pulse: {index} {total_pulses} * {index} = {pulse_list[index-1]}")
    
    return pulse_list

def get_total_pulses(distance, tire_size, total_miles, add_more_pulses=0):
    inches_distance = convert_feet_to_inch(distance)
    tire_rotation = calculate_tire_rotation_from_feet(inches_distance, tire_size)
    total_pulses = calculate_total_pulses(tire_rotation)
    return total_pulses 

def get_pulses_in_leg(distance, tire_size, total_miles, add_more_pulses=0):
    inches_distance = convert_feet_to_inch(distance)
    tire_rotation = calculate_tire_rotation_from_feet(inches_distance, tire_size)
    total_pulses = calculate_total_pulses(tire_rotation)
    # distance_size = len(mile_list)
    minimum_pulse = calculate_minimum_pulse(total_pulses, total_miles)
    # list_pulses = 
    # print("================================")
    # print(f"distance in feet = {distance:,}")
    # print(f"distance in inch = {inches_distance:,}")
    # print(f"tire rotation in inch = {tire_rotation:,}")
    # print(f"total tire rotation = {total_pulses:,}")
    # print(f"minimum pulse = {minimum_pulse}")
    total_miles = int(total_miles)
    print(f" = {total_miles}")
    pulses = calculate_list_pulses(minimum_pulse, total_miles, add_more_pulses)
    # print("================================")
    return pulses

# main application
def run(): 
    try:
        # a warning message 
        print("////////////////////////////////////////////////////")
        print("PLEASE, CLOSE ANY EXCEL FILE RELATED TO THIS PROGRAM")
        print("////////////////////////////////////////////////////\n")
        
        # in miles get the user data input 
        tire_size = 81.6
        total_length = 118.0
        my_speed = 130.0
        number_of_legs = 2
        # tire_size = float(input("Input the tire size (inches): "))
        # total_length = float(input("Input total race length (miles): "))
        # my_speed = float(input("Input target speed (mph): "))
        # number_of_legs = int(input("Input the number of legs (1 or 2) depending of the type of race: "))
        
        # distance_2st_leg = float(input("Input the end of 2nd leg (feet): "))
        
        # get mile and time list generated 
        mile_list = get_miles_list(total_length)
        timer_list = get_time_list(len(mile_list), my_speed)
        
        feet_distance = convert_miles_to_feet(total_length)
        inches_distance = convert_feet_to_inch(feet_distance)
        tire_rotation = calculate_tire_rotation_from_feet(inches_distance, tire_size)
        total_pulses = calculate_total_pulses(tire_rotation)
        distance_size = len(mile_list)
        minimum_pulse = calculate_minimum_pulse(total_pulses, distance_size)
        # list_pulses = 
        # print("================================")
        # print(f"distance in feet = {feet_distance}")
        # print(f"distance in inch = {inches_distance}")
        # print(f"tire rotation in inch = {tire_rotation}")
        # print(f"total tire rotation = {total_pulses}")
        # print(f"minimum pulse = {minimum_pulse}")
        
        # calculate_list_pulses(minimum_pulse, distance_size)
        # print(f"number of cells {len(mile_list)}")
        # calculate_list_pulses(total_pulses, mile_list)
        # print("================================")
        pulses = calculate_pulses(tire_size)
        
        if number_of_legs == 1:
            distance_1st_leg = float(input("Input the end of 1st leg (feet): "))

            # calculate pulses and get the pulses list
            p_list = get_pulses_list(pulses, len(mile_list))

            pulse_list = get_pulses_in_leg(distance_1st_leg, tire_size, len(mile_list)/2)
            total_pulses = get_total_pulses(distance_1st_leg, tire_size, len(mile_list)/2)

            # save the data into an spreadsheet
            save_list_in_excel(p_list, timer_list, mile_list, tire_size, total_length, my_speed, number_of_legs)
            print("\n////////////////////////////////////////////////////")
            print("SPREADSHEET CREATED, NOW THIS WINDOW IS ABLE TO BE CLOSED")
            print("////////////////////////////////////////////////////\n")
            input("Enter any key to exit")
        elif number_of_legs == 2:
            distance_1st_leg = 311221.0
            distance_2st_leg = 311256.0
            # distance_1st_leg = float(input("Input the end of 1st leg (feet): "))
            # distance_2st_leg = float(input("Input the end of 2nd leg (feet): "))

            print("Data for Leg 1: ")
            pulse_list = get_pulses_in_leg(distance_1st_leg, tire_size, len(mile_list)/2)
            total_pulses = get_total_pulses(distance_1st_leg, tire_size, len(mile_list)/2)
            
            # for i in pulse_list:
            #     print(i)
            pulse_list.remove(0)
            pulse_list.remove(pulse_list[-1])

            print("Data for Leg 2: ")
            pulse_list.extend(get_pulses_in_leg(distance_2st_leg, tire_size, len(mile_list)/2, total_pulses))
            total_pulses += get_total_pulses(distance_2st_leg, tire_size, len(mile_list)/2, total_pulses)
            for i, x in enumerate(pulse_list):
                print(i, x)
                
            # calculate pulses and get the pulses list
            # distance_end_1st_leg = float(input("Enter distance of the end of 1st leg (miles): "))
            # distance_end_1st_leg = get_pulse_by_distance(pulses, distance_end_1st_leg)
            # p_list = get_pulses_list(pulses, len(mile_list))

            p_list = pulse_list
            # for i in p_list:
            #     print(i)

            print(len(p_list))
            # save the data into an spreadsheet
            save_list_in_excel(p_list, timer_list, mile_list, tire_size, total_length, my_speed, number_of_legs, distance_1st_leg)
            print("\n////////////////////////////////////////////////////")
            print("SPREADSHEET CREATED, NOW THIS WINDOW IS ABLE TO BE CLOSED")
            print("////////////////////////////////////////////////////\n")
            input("Enter any key to exit")

        else:
            print("please Input a number of legs valid: (1 o 2)")
            exit()
    except PermissionError:
        print("There's a spreadsheet with the same name in the same location")
        input("choose another name and try again. \nPress any key to terminate this program.")


if __name__ == '__main__':
    # in miles get the user data input 
    tire_size = float(input("Input the tire size (inches): "))
    total_length = float(input("Input total race length (miles): "))
    my_speed = float(input("Input target speed (mph): "))
    # number_of_legs = int(input("Input the number of legs (1 or 2) depending of the type of race: "))
    distance_1st_leg = float(input("Input the end of 1st leg (feet): "))
    distance_2st_leg = float(input("Input the end of 2nd leg: "))
    
    pulse_list = []
    # get mile and time list generated 
    mile_list = get_miles_list(total_length)
    timer_list = get_time_list(len(mile_list), my_speed)
    print("Data for Leg 1: ")
    pulse_list = get_pulses_in_leg(distance_1st_leg, tire_size, len(mile_list)/2)
    total_pulses = get_total_pulses(distance_1st_leg, tire_size, len(mile_list)/2)
    print("Data for Leg 2: ")
    pulse_list.extend(get_pulses_in_leg(distance_2st_leg, tire_size, len(mile_list)/2, total_pulses))
    total_pulses += get_total_pulses(distance_2st_leg, tire_size, len(mile_list)/2, total_pulses)
    print(f"total pulses in the race: {total_pulses}")
    for i in pulse_list:
        print(i)

